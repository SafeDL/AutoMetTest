from torch.utils.data import Dataset
import os
import torchvision
import torchvision.transforms as transforms
from pytorch_fid import fid_score
from skimage.metrics import peak_signal_noise_ratio as psnr
from skimage.metrics import structural_similarity as ssim
import pandas as pd
import cv2
import openpyxl
import numpy as np
from VisionMetrics.inception_score import inception_metric
from PIL import Image
import argparse
import warnings

warnings.filterwarnings('ignore')


class CustomDataset(Dataset):
    def __init__(self, root_dir, transform=None):
        self.root_dir = root_dir
        self.images = os.listdir(root_dir)
        self.transform = transform

    def __len__(self):
        return len(self.images)

    def __getitem__(self, idx):
        img_name = os.path.join(self.root_dir, self.images[idx])
        image = Image.open(img_name)
        if self.transform:
            image = self.transform(image)

        return image


def scale_and_pad_image(img, target_size=(384, 640)):
    h, w = img.shape[:2]
    scale_ratio = min(target_size[0] / h, target_size[1] / w)
    new_size = (int(w * scale_ratio), int(h * scale_ratio))
    scaled_img = cv2.resize(img, new_size, interpolation=cv2.INTER_AREA)

    delta_w = target_size[1] - new_size[0]
    delta_h = target_size[0] - new_size[1]
    top, bottom = delta_h // 2, delta_h - (delta_h // 2)
    left, right = delta_w // 2, delta_w - (delta_w // 2)

    padded_img = cv2.copyMakeBorder(scaled_img, top, bottom, left, right, cv2.BORDER_CONSTANT, value=[0, 0, 0])
    return padded_img


def process_and_save_images(folder_path, target_size=(384, 640)):
    for filename in os.listdir(folder_path):
        img_path = os.path.join(folder_path, filename)
        img = cv2.imread(img_path)
        processed_img = scale_and_pad_image(img, target_size)
        processed_img_path = os.path.join(folder_path, filename)
        cv2.imwrite(processed_img_path, processed_img)

    return folder_path


def calculate_psnr_and_ssim(real_image, generated_image):
    real_image_np = np.array(real_image)
    generated_image_np = np.array(generated_image)

    psnr_value = psnr(real_image_np, generated_image_np)

    min_height, min_width = real_image_np.shape[:2] 
    win_size = min(7, min(min_height, min_width)) 

    if win_size < 1:  
        ssim_value = float('nan')  
    else:
        if win_size % 2 == 0: 
            win_size -= 1
        ssim_value = ssim(real_image_np, generated_image_np, channel_axis=-1, win_size=win_size)

    return psnr_value, ssim_value


if __name__ == '__main__':
    # Configure argument parser
    parser = argparse.ArgumentParser(description='Select the appropriate mutation method')
    parser.add_argument('--search_type', type=str, default='CycleGAN',
                        help='TYPE:TACTIC/MUNIT/DeepRoad/DeepTest/CycleGAN')
    args = parser.parse_args()

    # Load the pre-trained Inception-v3 model
    inception_model = torchvision.models.inception_v3(pretrained=True)

    # Define preprocessing transforms
    transform = transforms.Compose([
        transforms.Resize((384, 640)),
        transforms.ToTensor(),
        transforms.Normalize(mean=[0.5, 0.5, 0.5], std=[0.5, 0.5, 0.5])
    ])

    # Mutated environmental types (DeepRoad/TACTIC/MUNIT/CycleGAN)
    mutate_conditions = ['cloudy', 'dawndusk', 'overcast', 'foggy', 'rainy', 'night', 'night_rainy', 'night_fog']

    # Just for DeepTest
    # mutate_conditions = ['image_translation', 'image_shear', 'image_scale', 'image_rotation',
    #                      'image_contrast', 'image_brightness', 'image_blur', 'image_add_noise']

    original_conditions = [1, 3, 45, 69, 70, 71, 72, 75, 76, 77, 80, 81, 82, 83, 85, 86, 87, 88,
                           89, 90, 95, 98, 115, 120, 123, 127, 130, 131, 133, 134, 136, 165]

    # Specify the path to store results
    excel_path = os.path.join('../Metamorphic/mutation_results.xlsx')
    workbook = openpyxl.load_workbook(excel_path)
    if args.search_type == 'TACTIC':
        worksheet = workbook.worksheets[8]
    elif args.search_type == 'MUNIT':
        worksheet = workbook.worksheets[9]
    elif args.search_type == 'DeepRoad':
        worksheet = workbook.worksheets[10]
    elif args.search_type == 'DeepTest':
        worksheet = workbook.worksheets[11]
    elif args.search_type == 'CycleGAN':
        worksheet = workbook.worksheets[12]
    else:
        raise ValueError("search type cannot be found")

    # Calculate image mutation quality (DeepRoad/TACTIC/MUNIT/CycleGAN)
    for mutate_condition in mutate_conditions:
        if mutate_condition == 'night':
            fid_score_column = 4
            inception_score_column = 5
            PSNR_column = 6
            SSIM_column = 7
        elif mutate_condition == 'overcast':
            fid_score_column = 10
            inception_score_column = 11
            PSNR_column = 12
            SSIM_column = 13
        elif mutate_condition == 'rainy':
            fid_score_column = 16
            inception_score_column = 17
            PSNR_column = 18
            SSIM_column = 19
        elif mutate_condition == 'foggy':
            fid_score_column = 22
            inception_score_column = 23
            PSNR_column = 24
            SSIM_column = 25
        elif mutate_condition == 'dawndusk':
            fid_score_column = 28
            inception_score_column = 29
            PSNR_column = 30
            SSIM_column = 31
        elif mutate_condition == 'cloudy':
            fid_score_column = 34
            inception_score_column = 35
            PSNR_column = 36
            SSIM_column = 37
        elif mutate_condition == 'night_rainy':
            fid_score_column = 40
            inception_score_column = 41
            PSNR_column = 42
            SSIM_column = 43
        elif mutate_condition == 'night_fog':
            fid_score_column = 46
            inception_score_column = 47
            PSNR_column = 48
            SSIM_column = 49
        else:
            raise ValueError("mutate condition must be one of cloudy, dawndusk, overcast, foggy, rainy, night")

    # Calculate image mutation quality (just for DeepTest)
    # for mutate_condition in mutate_conditions:
    #     if mutate_condition == 'image_translation':
    #         fid_score_column = 4
    #         inception_score_column = 5
    #         PSNR_column = 6
    #         SSIM_column = 7
    #     elif mutate_condition == 'image_scale':
    #         fid_score_column = 10
    #         inception_score_column = 11
    #         PSNR_column = 12
    #         SSIM_column = 13
    #     elif mutate_condition == 'image_shear':
    #         fid_score_column = 16
    #         inception_score_column = 17
    #         PSNR_column = 18
    #         SSIM_column = 19
    #     elif mutate_condition == 'image_rotation':
    #         fid_score_column = 22
    #         inception_score_column = 23
    #         PSNR_column = 24
    #         SSIM_column = 25
    #     elif mutate_condition == 'image_contrast':
    #         fid_score_column = 28
    #         inception_score_column = 29
    #         PSNR_column = 30
    #         SSIM_column = 31
    #     elif mutate_condition == 'image_brightness':
    #         fid_score_column = 34
    #         inception_score_column = 35
    #         PSNR_column = 36
    #         SSIM_column = 37
    #     elif mutate_condition == 'image_blur':
    #         fid_score_column = 40
    #         inception_score_column = 41
    #         PSNR_column = 42
    #         SSIM_column = 43
    #     elif mutate_condition == 'image_add_noise':
    #         fid_score_column = 46
    #         inception_score_column = 47
    #         PSNR_column = 48
    #         SSIM_column = 49
    #     else:
    #         raise ValueError("mutate condition must be one of cloudy, dawndusk, overcast, foggy, rainy, night")

        for idx, condition in enumerate(original_conditions):
            name = 'condition_%d' % condition
            print(f"we are now testing mutated {name}")
            real_images_folder = os.path.join('../Autopilot/driving_dataset/carla_collect', name, 'images')
            if args.search_type == 'TACTIC':
                generated_images_folder = os.path.join(
                    './TACTIC/test_outputs/%s_tactic/condition_%d' % (mutate_condition, condition))
            elif args.search_type == 'MUNIT':
                generated_images_folder = os.path.join(
                    './TACTIC/test_outputs/%s_munit/condition_%d' % (mutate_condition, condition))
            elif args.search_type == 'DeepRoad':
                generated_images_folder = os.path.join(
                    './TACTIC/test_outputs/%s_deeproad/condition_%d' % (mutate_condition, condition))
            elif args.search_type == 'CycleGAN':
                generated_images_folder = os.path.join(
                    './CycleGAN/%s/condition_%d' % (mutate_condition, condition))
            elif args.search_type == 'DeepTest':
                generated_images_folder = os.path.join('./DeepTest/test_outputs/%s/condition_%d' % (mutate_condition, condition))
                if mutate_condition == 'image_scale':
                    process_and_save_images(generated_images_folder)
            else:
                raise ValueError("search type must be random or ea or unit")

            # Calculate Inception Score and FID
            dataset = CustomDataset(root_dir=generated_images_folder, transform=transform)
            inception_score = inception_metric(dataset, cuda=True, batch_size=32, resize=True, splits=10)
            fid_value = fid_score.calculate_fid_given_paths([real_images_folder, generated_images_folder],
                                                            batch_size=32, device='cuda', dims=2048)
            print(f"the FID of {name} is {fid_value}")
            print(f"the IS of {name} is {inception_score[0]}")

            # Calculate PSNR and SSIM
            psnr_values = []
            ssim_values = []
            for img_name in os.listdir(real_images_folder):
                real_image = Image.open(os.path.join(real_images_folder, img_name))
                generated_image = Image.open(os.path.join(generated_images_folder, img_name))

                # Calculate PSNR and SSIM
                psnr_value, ssim_value = calculate_psnr_and_ssim(real_image, generated_image)
                psnr_values.append(psnr_value)
                ssim_values.append(ssim_value)

            # Calculate average PSNR and SSIM
            avg_psnr = np.mean(psnr_values)
            avg_ssim = np.mean(ssim_values)

            # Write the statistics to the excel file
            worksheet.cell(idx + 2, column=fid_score_column, value=fid_value)
            worksheet.cell(idx + 2, column=inception_score_column, value=float(inception_score[0]))
            worksheet.cell(idx + 2, column=PSNR_column, value=avg_psnr)
            worksheet.cell(idx + 2, column=SSIM_column, value=avg_ssim)
            workbook.save(excel_path)
