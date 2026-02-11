from torch.utils.data import Dataset
import os
import torchvision
import torchvision.transforms as transforms
from pytorch_fid import fid_score
import numpy as np
from skimage.metrics import peak_signal_noise_ratio as psnr
from skimage.metrics import structural_similarity as ssim
import openpyxl
from VisionMetrics.inception_score import inception_metric
from PIL import Image
import warnings

warnings.filterwarnings('ignore')


class CustomDataset(Dataset):
    def __init__(self, root_dir, transform=None):
        self.root_dir = root_dir
        self.images = os.listdir(root_dir)
        self.transform = transform

    def __len__(self):
        return len(self.images)

    def __getitem__(self, idx):
        img_name = os.path.join(self.root_dir, self.images[idx])
        image = Image.open(img_name)

        if self.transform:
            image = self.transform(image)

        return image


def calculate_psnr_and_ssim(real_image, generated_image):
    real_image_np = np.array(real_image)
    generated_image_np = np.array(generated_image)

    # Calculate PSNR
    psnr_value = psnr(real_image_np, generated_image_np)

    # Calculate SSIM, dynamically adjust win_size
    min_height, min_width = real_image_np.shape[:2] 
    win_size = min(7, min(min_height, min_width)) 

    if win_size < 1:  
        ssim_value = float('nan') 
    else:
        if win_size % 2 == 0:  
            win_size -= 1
    ssim_value = ssim(real_image_np, generated_image_np, channel_axis=-1, win_size=win_size)

    return psnr_value, ssim_value


if __name__ == '__main__':
    # Load pretrained Inception-v3 model
    inception_model = torchvision.models.inception_v3(pretrained=True)

    # Define preprocessing transforms
    transform = transforms.Compose([
        transforms.Resize((384, 640)),
        transforms.ToTensor(),
        transforms.Normalize(mean=[0.5, 0.5, 0.5], std=[0.5, 0.5, 0.5])
    ])
    # Define the environmental conditions to mutate
    original_conditions = [1, 3, 45, 69, 70, 71, 72, 75, 76, 77, 80, 81, 82, 83, 85, 86, 87, 88,
                           89, 90, 95, 98, 115, 120, 123, 127, 130, 131, 133, 134, 136, 165]
    for idx in range(0, 8): 
        # Specify the path to store results
        excel_path = os.path.join('./mutation_results.xlsx')
        workbook = openpyxl.load_workbook(excel_path)
        worksheet = workbook.worksheets[idx]  

        for jdx in range(0, len(original_conditions)):
            name = 'condition_%d' % original_conditions[jdx]
            print(f"Processing {name} in mutation_{idx}")
            real_images_folder = os.path.join('../Autopilot/driving_dataset/carla_collect', name, 'images')
            generated_images_folder = os.path.join('mutation_%d' % idx, name)

            # Build custom dataset for FID and Inception calculation
            dataset = CustomDataset(root_dir=generated_images_folder, transform=transform)
            inception_score = inception_metric(dataset, cuda=True, batch_size=32, resize=True, splits=10)
            fid_value = fid_score.calculate_fid_given_paths([real_images_folder, generated_images_folder],
                                                            batch_size=32, device='cuda', dims=2048)

            # Calculate PSNR and SSIM
            psnr_values = []
            ssim_values = []
            for img_name in os.listdir(real_images_folder):
                real_image = Image.open(os.path.join(real_images_folder, img_name))
                generated_image = Image.open(os.path.join(generated_images_folder, img_name))

                psnr_value, ssim_value = calculate_psnr_and_ssim(real_image, generated_image)
                psnr_values.append(psnr_value)
                ssim_values.append(ssim_value)

            # Calculate average PSNR and SSIM
            avg_psnr = np.mean(psnr_values)
            avg_ssim = np.mean(ssim_values)

            # Write statistics to the excel file
            worksheet.cell(jdx + 2, column=12, value=fid_value)
            worksheet.cell(jdx + 2, column=13, value=float(inception_score[0]))
            worksheet.cell(jdx + 2, column=14, value=avg_psnr)  
            worksheet.cell(jdx + 2, column=15, value=avg_ssim) 
            workbook.save(excel_path)
